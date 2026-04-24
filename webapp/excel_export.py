"""Prüfer-Excel-Export: IDV-Register, Maßnahmen, Prüfungen, Nachweise.

Erzeugt eine mehrseitige Arbeitsmappe mit Deckblatt und bedingter
Formatierung (Ampel für Fristen). Wird von ``admin.export_excel`` aufgerufen.
"""
from __future__ import annotations

import io
import sqlite3
from datetime import date, datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .routes.eigenentwicklung import ENTWICKLUNGSART_LABEL


# ---------------------------------------------------------------------------
# Formatierungs-Helfer
# ---------------------------------------------------------------------------

_HEADER_FILL  = PatternFill("solid", fgColor="1F3A5F")
_HEADER_FONT  = Font(color="FFFFFF", bold=True)
_TITLE_FONT   = Font(size=16, bold=True, color="1F3A5F")

_AMPEL_ROT    = PatternFill("solid", fgColor="F8D7DA")
_AMPEL_GELB   = PatternFill("solid", fgColor="FFF3CD")
_AMPEL_GRUEN  = PatternFill("solid", fgColor="D1E7DD")
_AMPEL_GRAU   = PatternFill("solid", fgColor="E9ECEF")


def _parse_iso(value: Optional[str]) -> Optional[date]:
    if not value:
        return None
    try:
        return date.fromisoformat(value[:10])
    except ValueError:
        return None


def _faelligkeit_fill(faellig: Optional[str], erledigt: bool) -> Optional[PatternFill]:
    """Ampel-Fill je nach Fälligkeits-/Erledigungs-Status."""
    if erledigt:
        return _AMPEL_GRAU
    d = _parse_iso(faellig)
    if not d:
        return None
    heute = date.today()
    delta = (d - heute).days
    if delta < 0:
        return _AMPEL_ROT
    if delta <= 30:
        return _AMPEL_GELB
    return _AMPEL_GRUEN


def _write_header(ws, labels: list[tuple[str, int]], row: int = 1) -> None:
    for col_idx, (label, width) in enumerate(labels, start=1):
        cell = ws.cell(row=row, column=col_idx, value=label)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center", horizontal="left")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[row].height = 22
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def _autofilter(ws, n_cols: int, n_rows: int) -> None:
    if n_rows < 1:
        return
    ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}{n_rows + 1}"


# ---------------------------------------------------------------------------
# Sheet-Builder
# ---------------------------------------------------------------------------

def _sheet_deckblatt(wb: Workbook, db: sqlite3.Connection) -> None:
    ws = wb.active
    ws.title = "Deckblatt"

    ws["A1"] = "IDV-Register – Prüfer-Export"
    ws["A1"].font = _TITLE_FONT
    ws["A2"] = f"Stichtag: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(italic=True, color="475569")

    aktive_idvs = db.execute(
        "SELECT COUNT(*) FROM idv_register WHERE status NOT IN ('Außer Betrieb','Abgelöst')"
    ).fetchone()[0]
    wesentliche = db.execute("""
        SELECT COUNT(*) FROM idv_register r
         WHERE r.status NOT IN ('Außer Betrieb','Abgelöst')
           AND EXISTS (SELECT 1 FROM idv_wesentlichkeit w
                        JOIN wesentlichkeitskriterien k ON k.id = w.kriterium_id
                        WHERE w.idv_db_id = r.id AND w.erfuellt = 1 AND k.aktiv = 1)
    """).fetchone()[0]
    massnahmen_offen = db.execute(
        "SELECT COUNT(*) FROM massnahmen WHERE status IN ('Offen','In Bearbeitung')"
    ).fetchone()[0]
    massnahmen_ueberfaellig = db.execute(
        "SELECT COUNT(*) FROM massnahmen"
        " WHERE status IN ('Offen','In Bearbeitung') AND faellig_am < date('now')"
    ).fetchone()[0]
    pruefungen_faellig = db.execute("""
        SELECT COUNT(*) FROM idv_register
         WHERE status NOT IN ('Außer Betrieb','Abgelöst')
           AND naechste_pruefung IS NOT NULL
           AND naechste_pruefung < date('now','+30 day')
    """).fetchone()[0]

    rows = [
        ("Aktive IDVs",                   aktive_idvs),
        ("Davon wesentlich",              wesentliche),
        ("Offene Maßnahmen",              massnahmen_offen),
        ("Davon überfällig",              massnahmen_ueberfaellig),
        ("Prüfungen fällig (≤30 Tage)",   pruefungen_faellig),
    ]
    for i, (label, value) in enumerate(rows, start=4):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=value)

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 14

    ws["A11"] = ("Arbeitsmappe enthält: Register, Maßnahmen, Prüfungen, Nachweise (Freigaben). "
                 "Farben in den Fristen-Spalten: rot = überfällig, gelb = fällig innerhalb 30 Tagen, "
                 "grün = mehr als 30 Tage, grau = erledigt/abgeschlossen.")
    ws["A11"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A11:D14")


def _sheet_register(wb: Workbook, db: sqlite3.Connection) -> None:
    ws = wb.create_sheet("Register")
    header = [
        ("IDV-ID",                 14),
        ("Bezeichnung",            38),
        ("Status",                 14),
        ("Entwicklungsart",        18),
        ("IDV-Typ",                18),
        ("Wesentlich",             12),
        ("Organisationseinheit",   24),
        ("Fachverantwortlicher",   24),
        ("Entwickler",             24),
        ("Koordinator",            24),
        ("Version",                10),
        ("Produktiv seit",         14),
        ("Letzte Prüfung",         16),
        ("Nächste Prüfung",        16),
        ("Prüfintervall (Mon.)",   18),
        ("Freigabe-Verfahren",     20),
    ]
    _write_header(ws, header)

    rows = db.execute("""
        SELECT r.idv_id, r.bezeichnung, r.status, r.entwicklungsart, r.idv_typ,
               EXISTS (
                   SELECT 1 FROM idv_wesentlichkeit w
                   JOIN wesentlichkeitskriterien k ON k.id = w.kriterium_id
                   WHERE w.idv_db_id = r.id AND w.erfuellt = 1 AND k.aktiv = 1
               ) AS wesentlich,
               ou.bezeichnung AS oe,
               (pf.nachname || ', ' || pf.vorname)   AS fv,
               (pe.nachname || ', ' || pe.vorname)   AS entwickler,
               (pk.nachname || ', ' || pk.vorname)   AS koordinator,
               r.version, r.produktiv_seit,
               (SELECT MAX(pruefungsdatum) FROM pruefungen p WHERE p.idv_id = r.id
                  AND p.abgeschlossen = 1) AS letzte_pruefung,
               r.naechste_pruefung, r.pruefintervall_monate,
               COALESCE(r.freigabe_verfahren, 'Standard') AS freigabe_verfahren
          FROM idv_register r
          LEFT JOIN org_units ou ON ou.id = r.org_unit_id
          LEFT JOIN persons  pf  ON pf.id = r.fachverantwortlicher_id
          LEFT JOIN persons  pe  ON pe.id = r.idv_entwickler_id
          LEFT JOIN persons  pk  ON pk.id = r.idv_koordinator_id
         ORDER BY r.idv_id
    """).fetchall()

    for i, r in enumerate(rows, start=2):
        ws.cell(row=i, column=1,  value=r["idv_id"])
        ws.cell(row=i, column=2,  value=r["bezeichnung"])
        ws.cell(row=i, column=3,  value=r["status"])
        ws.cell(row=i, column=4,  value=ENTWICKLUNGSART_LABEL.get(
            r["entwicklungsart"], r["entwicklungsart"]))
        ws.cell(row=i, column=5,  value=r["idv_typ"])
        ws.cell(row=i, column=6,  value="Ja" if r["wesentlich"] else "")
        ws.cell(row=i, column=7,  value=r["oe"])
        ws.cell(row=i, column=8,  value=r["fv"])
        ws.cell(row=i, column=9,  value=r["entwickler"])
        ws.cell(row=i, column=10, value=r["koordinator"])
        ws.cell(row=i, column=11, value=r["version"])
        ws.cell(row=i, column=12, value=r["produktiv_seit"])
        ws.cell(row=i, column=13, value=r["letzte_pruefung"])
        c_faelligkeit = ws.cell(row=i, column=14, value=r["naechste_pruefung"])
        fill = _faelligkeit_fill(r["naechste_pruefung"], erledigt=False)
        if fill:
            c_faelligkeit.fill = fill
        ws.cell(row=i, column=15, value=r["pruefintervall_monate"])
        ws.cell(row=i, column=16, value=r["freigabe_verfahren"])

    _autofilter(ws, len(header), len(rows))


def _sheet_massnahmen(wb: Workbook, db: sqlite3.Connection) -> None:
    ws = wb.create_sheet("Maßnahmen")
    header = [
        ("IDV-ID",              14),
        ("IDV-Bezeichnung",     32),
        ("Titel",               38),
        ("Typ",                 18),
        ("Priorität",           12),
        ("Status",              16),
        ("Verantwortlicher",    24),
        ("Fällig am",           14),
        ("Erledigt am",         14),
    ]
    _write_header(ws, header)

    rows = db.execute("""
        SELECT r.idv_id, r.bezeichnung,
               m.titel, m.massnahmentyp, m.prioritaet, m.status,
               (p.nachname || ', ' || p.vorname) AS verantwortlicher,
               m.faellig_am, m.erledigt_am
          FROM massnahmen m
          JOIN idv_register r ON r.id = m.idv_id
          LEFT JOIN persons p ON p.id = m.verantwortlicher_id
         ORDER BY m.faellig_am IS NULL, m.faellig_am, r.idv_id
    """).fetchall()

    for i, r in enumerate(rows, start=2):
        erledigt = r["status"] == "Erledigt"
        ws.cell(row=i, column=1, value=r["idv_id"])
        ws.cell(row=i, column=2, value=r["bezeichnung"])
        ws.cell(row=i, column=3, value=r["titel"])
        ws.cell(row=i, column=4, value=r["massnahmentyp"])
        ws.cell(row=i, column=5, value=r["prioritaet"])
        ws.cell(row=i, column=6, value=r["status"])
        ws.cell(row=i, column=7, value=r["verantwortlicher"])
        c_faellig = ws.cell(row=i, column=8, value=r["faellig_am"])
        fill = _faelligkeit_fill(r["faellig_am"], erledigt=erledigt)
        if fill:
            c_faellig.fill = fill
        ws.cell(row=i, column=9, value=r["erledigt_am"])

    _autofilter(ws, len(header), len(rows))


def _sheet_pruefungen(wb: Workbook, db: sqlite3.Connection) -> None:
    ws = wb.create_sheet("Prüfungen")
    header = [
        ("IDV-ID",             14),
        ("IDV-Bezeichnung",    32),
        ("Prüfungsart",        18),
        ("Prüfungsdatum",      14),
        ("Prüfer",             24),
        ("Ergebnis",           18),
        ("Maßn. erforderlich", 18),
        ("Frist Maßnahmen",    16),
        ("Abgeschlossen",      14),
        ("Nächste Prüfung",    16),
    ]
    _write_header(ws, header)

    rows = db.execute("""
        SELECT r.idv_id, r.bezeichnung,
               p.pruefungsart, p.pruefungsdatum,
               (pe.nachname || ', ' || pe.vorname) AS pruefer,
               p.ergebnis, p.massnahmen_erforderlich, p.frist_massnahmen,
               p.abgeschlossen, p.naechste_pruefung
          FROM pruefungen p
          JOIN idv_register r ON r.id = p.idv_id
          LEFT JOIN persons  pe ON pe.id = p.pruefer_id
         ORDER BY p.pruefungsdatum DESC, r.idv_id
    """).fetchall()

    for i, r in enumerate(rows, start=2):
        abgeschlossen = bool(r["abgeschlossen"])
        ws.cell(row=i, column=1,  value=r["idv_id"])
        ws.cell(row=i, column=2,  value=r["bezeichnung"])
        ws.cell(row=i, column=3,  value=r["pruefungsart"])
        ws.cell(row=i, column=4,  value=r["pruefungsdatum"])
        ws.cell(row=i, column=5,  value=r["pruefer"])
        c_erg = ws.cell(row=i, column=6, value=r["ergebnis"])
        if (r["ergebnis"] or "").startswith("Kritisch") or r["ergebnis"] == "Nicht bestanden":
            c_erg.fill = _AMPEL_ROT
        ws.cell(row=i, column=7,  value="Ja" if r["massnahmen_erforderlich"] else "Nein")
        c_frist = ws.cell(row=i, column=8, value=r["frist_massnahmen"])
        fill = _faelligkeit_fill(r["frist_massnahmen"], erledigt=abgeschlossen)
        if fill:
            c_frist.fill = fill
        ws.cell(row=i, column=9,  value="Ja" if abgeschlossen else "Nein")
        ws.cell(row=i, column=10, value=r["naechste_pruefung"])

    _autofilter(ws, len(header), len(rows))


def _sheet_nachweise(wb: Workbook, db: sqlite3.Connection) -> None:
    ws = wb.create_sheet("Nachweise")
    header = [
        ("IDV-ID",           14),
        ("IDV-Bezeichnung",  32),
        ("Schritt",          30),
        ("Status",           16),
        ("Beauftragt am",    16),
        ("Beauftragt von",   22),
        ("Zugewiesen an",    22),
        ("Durchgeführt am",  16),
        ("Durchgeführt von", 22),
        ("Archiv-SHA256",    24),
    ]
    _write_header(ws, header)

    rows = db.execute("""
        SELECT r.idv_id, r.bezeichnung,
               f.schritt, f.status,
               f.beauftragt_am,
               (pb.nachname || ', ' || pb.vorname) AS beauftragt_von,
               (pz.nachname || ', ' || pz.vorname) AS zugewiesen_an,
               f.durchgefuehrt_am,
               (pd.nachname || ', ' || pd.vorname) AS durchgefuehrt_von,
               f.archiv_datei_sha256
          FROM idv_freigaben f
          JOIN idv_register r  ON r.id = f.idv_id
          LEFT JOIN persons pb ON pb.id = f.beauftragt_von_id
          LEFT JOIN persons pz ON pz.id = f.zugewiesen_an_id
          LEFT JOIN persons pd ON pd.id = f.durchgefuehrt_von_id
         ORDER BY r.idv_id, f.beauftragt_am
    """).fetchall()

    for i, r in enumerate(rows, start=2):
        ws.cell(row=i, column=1,  value=r["idv_id"])
        ws.cell(row=i, column=2,  value=r["bezeichnung"])
        ws.cell(row=i, column=3,  value=r["schritt"])
        c_status = ws.cell(row=i, column=4, value=r["status"])
        if r["status"] == "Erledigt":
            c_status.fill = _AMPEL_GRUEN
        elif r["status"] in ("Nicht erledigt", "Abgebrochen"):
            c_status.fill = _AMPEL_ROT
        elif r["status"] == "Ausstehend":
            c_status.fill = _AMPEL_GELB
        ws.cell(row=i, column=5,  value=(r["beauftragt_am"] or "")[:10])
        ws.cell(row=i, column=6,  value=r["beauftragt_von"])
        ws.cell(row=i, column=7,  value=r["zugewiesen_an"])
        ws.cell(row=i, column=8,  value=(r["durchgefuehrt_am"] or "")[:10])
        ws.cell(row=i, column=9,  value=r["durchgefuehrt_von"])
        ws.cell(row=i, column=10, value=r["archiv_datei_sha256"])

    _autofilter(ws, len(header), len(rows))


def _sheet_kennzahlen(wb: Workbook, db: sqlite3.Connection) -> None:
    """Issue #354: Prozesskennzahlen fuer Revision/Aufsicht.

    Ein Sheet ``Kennzahlen`` mit den Werten fuer 30 und 90 Tage. Quelle
    ist ``db.get_dashboard_kpis``; die Werte basieren ausschliesslich auf
    Audit-Trail / Notification-Log / Self-Service-Audit (kein neuer
    DB-State).
    """
    import sys, os
    sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    from db import get_dashboard_kpis  # noqa: WPS433

    ws = wb.create_sheet("Kennzahlen")
    header = [
        ("Kennzahl",        38),
        ("Wert (30 Tage)",  18),
        ("Detail (30 Tage)",36),
        ("Wert (90 Tage)",  18),
        ("Detail (90 Tage)",36),
    ]
    _write_header(ws, header)

    kpis_30 = get_dashboard_kpis(db, days=30)
    kpis_90 = get_dashboard_kpis(db, days=90)
    by_key_90 = {k["key"]: k for k in kpis_90}

    for i, k in enumerate(kpis_30, start=2):
        k90 = by_key_90.get(k["key"], {})
        ws.cell(row=i, column=1, value=k["label"])
        ws.cell(row=i, column=2, value=k["value"])
        ws.cell(row=i, column=3, value=k["sub"])
        ws.cell(row=i, column=4, value=k90.get("value", ""))
        ws.cell(row=i, column=5, value=k90.get("sub", ""))

    _autofilter(ws, len(header), len(kpis_30))


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def build_register_workbook(db: sqlite3.Connection) -> Workbook:
    """Baut die vollständige Prüfer-Arbeitsmappe auf."""
    wb = Workbook()
    _sheet_deckblatt(wb, db)
    _sheet_register(wb, db)
    _sheet_massnahmen(wb, db)
    _sheet_pruefungen(wb, db)
    _sheet_nachweise(wb, db)
    _sheet_kennzahlen(wb, db)
    return wb


def register_excel_bytes(db: sqlite3.Connection) -> bytes:
    """Serialisiert die Arbeitsmappe für ``send_file`` bzw. Response-Body."""
    wb = build_register_workbook(db)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Report: Excel-Dateien ohne Zell-/Blattschutz (IT-Risiko-Kandidaten)
# ---------------------------------------------------------------------------

_UNPROTECTED_EXCEL_EXTS = (".xlsx", ".xlsm", ".xlsb", ".xltm", ".xltx")


def _sheet_ohne_schutz_deckblatt(wb: Workbook, anzahl: int) -> None:
    ws = wb.active
    ws.title = "Deckblatt"

    ws["A1"] = "Excel-Dateien ohne Zell-/Blattschutz"
    ws["A1"].font = _TITLE_FONT
    ws["A2"] = f"Stichtag: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(italic=True, color="475569")

    ws["A4"] = "Anzahl betroffener Dateien:"
    ws["A4"].font = Font(bold=True)
    ws.cell(row=4, column=2, value=anzahl).font = Font(bold=True, color="B91C1C")

    ws["A6"] = (
        "Aufsichtsrechtlicher Hintergrund: MaRisk AT 7.2 / DORA fordern, dass "
        "Individuelle Datenverarbeitung (IDV) vor unbeabsichtigten Änderungen "
        "geschützt ist. Für Excel-Tabellen umfasst dies insbesondere den "
        "Schutz von Formelzellen und Eingabemasken über Blatt- bzw. "
        "Arbeitsmappenschutz."
    )
    ws["A6"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A6:F10")

    ws["A12"] = (
        "Quelle: Scanner-Analyse der OOXML-Container (xl/workbook.xml + "
        "xl/worksheets/sheet*.xml). Berücksichtigt werden .xlsx, .xlsm, "
        ".xlsb, .xltm, .xltx. Binäre .xls-Dateien lassen sich nicht prüfen "
        "und sind nicht enthalten."
    )
    ws["A12"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A12"].font = Font(italic=True, color="475569")
    ws.merge_cells("A12:F15")

    ws["A17"] = (
        "Hinweis: Die Liste ist als Arbeitsgrundlage für die Anlage von "
        "IT-Risiken im IDV-Register gedacht. Pro Eintrag ist zu bewerten, "
        "ob ein Zellschutz zwingend erforderlich ist oder eine begründete "
        "Ausnahme dokumentiert werden kann. Die bewusste Akzeptanz einer "
        "fehlenden Protektion wird vom Fachverantwortlichen im Rahmen der "
        "Fachlichen Abnahme (Phase 2) gesetzt – siehe Spalten „Akzeptiert "
        "von / am / Begründung“. Grüne Zelle = akzeptiert, rote Zelle = "
        "einer IDV zugeordnet, aber Akzeptanz noch offen."
    )
    ws["A17"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A17:F22")

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 14


def _sheet_ohne_schutz_liste(wb: Workbook, db: sqlite3.Connection) -> list:
    ws = wb.create_sheet("Ohne Zellschutz")
    header = [
        ("Dateiname",              38),
        ("Vollständiger Pfad",     70),
        ("Share-Root",             28),
        ("Endung",                 10),
        ("Blätter",                10),
        ("Formelzellen",           14),
        ("Makros (VBA)",           14),
        ("Externe Verknüpfungen",  20),
        ("Office-Autor",           22),
        ("Datei-Eigentümer",       22),
        ("Zuletzt geändert",       20),
        ("Zuletzt gesehen",        20),
        ("Bearbeitungsstatus",     20),
        ("IDV-ID (falls verknüpft)", 18),
        ("Akzeptiert von",         24),
        ("Akzeptiert am",          16),
        ("Begründung (Akzeptanz)", 48),
    ]
    _write_header(ws, header)

    placeholders = ",".join("?" * len(_UNPROTECTED_EXCEL_EXTS))
    rows = db.execute(f"""
        SELECT f.file_name, f.full_path, f.share_root, f.extension,
               f.sheet_count, f.formula_count, f.has_macros,
               f.has_external_links, f.office_author, f.file_owner,
               f.modified_at, f.last_seen_at, f.bearbeitungsstatus,
               COALESCE(reg.idv_id, lnk_reg.idv_id)  AS idv_id,
               COALESCE(reg.id,     lnk_reg.id)      AS idv_db_id,
               az.akzeptiert_am,
               az.begruendung,
               (p.nachname || ', ' || p.vorname) AS akzeptiert_von
          FROM idv_files f
          LEFT JOIN idv_register  reg     ON reg.file_id = f.id
          LEFT JOIN idv_file_links lnk    ON lnk.file_id = f.id
          LEFT JOIN idv_register  lnk_reg ON lnk_reg.id  = lnk.idv_db_id
          LEFT JOIN idv_zellschutz_akzeptanz az
                 ON az.file_id = f.id
                AND az.idv_db_id = COALESCE(reg.id, lnk_reg.id)
          LEFT JOIN persons p ON p.id = az.akzeptiert_von_id
         WHERE f.status = 'active'
           AND LOWER(f.extension) IN ({placeholders})
           AND COALESCE(f.has_sheet_protection, 0) = 0
           AND COALESCE(f.workbook_protected, 0) = 0
           AND (f.bearbeitungsstatus IS NULL OR f.bearbeitungsstatus != 'Ignoriert')
         ORDER BY f.share_root, f.full_path
    """, _UNPROTECTED_EXCEL_EXTS).fetchall()

    for i, r in enumerate(rows, start=2):
        ws.cell(row=i, column=1,  value=r["file_name"])
        ws.cell(row=i, column=2,  value=r["full_path"])
        ws.cell(row=i, column=3,  value=r["share_root"])
        ws.cell(row=i, column=4,  value=r["extension"])
        ws.cell(row=i, column=5,  value=r["sheet_count"])
        ws.cell(row=i, column=6,  value=r["formula_count"])
        c_vba = ws.cell(row=i, column=7, value="Ja" if r["has_macros"] else "Nein")
        if r["has_macros"]:
            c_vba.fill = _AMPEL_ROT
        ws.cell(row=i, column=8,  value="Ja" if r["has_external_links"] else "Nein")
        ws.cell(row=i, column=9,  value=r["office_author"])
        ws.cell(row=i, column=10, value=r["file_owner"])
        ws.cell(row=i, column=11, value=(r["modified_at"] or "")[:19].replace("T", " "))
        ws.cell(row=i, column=12, value=(r["last_seen_at"] or "")[:19].replace("T", " "))
        ws.cell(row=i, column=13, value=r["bearbeitungsstatus"])
        ws.cell(row=i, column=14, value=r["idv_id"])
        ws.cell(row=i, column=15, value=r["akzeptiert_von"])
        c_akz_am = ws.cell(
            row=i, column=16,
            value=(r["akzeptiert_am"] or "")[:10],
        )
        if r["akzeptiert_am"]:
            c_akz_am.fill = _AMPEL_GRUEN
        elif r["idv_db_id"]:
            # Datei ist einer IDV zugeordnet, aber Akzeptanz fehlt → offenes IT-Risiko
            c_akz_am.fill = _AMPEL_ROT
        ws.cell(row=i, column=17, value=r["begruendung"])

    _autofilter(ws, len(header), len(rows))
    return rows


def build_unprotected_excel_workbook(db: sqlite3.Connection) -> Workbook:
    """Arbeitsmappe mit Excel-Dateien ohne Blatt-/Arbeitsmappenschutz."""
    wb = Workbook()
    # Deckblatt übernimmt das automatisch angelegte Default-Sheet
    # (analog zu build_register_workbook); die Anzahl wird nachträglich gesetzt.
    rows_count_placeholder = 0
    _sheet_ohne_schutz_deckblatt(wb, anzahl=rows_count_placeholder)
    rows = _sheet_ohne_schutz_liste(wb, db)
    # Anzahl nachtragen, jetzt wo wir sie kennen
    wb["Deckblatt"].cell(row=4, column=2, value=len(rows)).font = Font(
        bold=True, color="B91C1C"
    )
    return wb


def unprotected_excel_bytes(db: sqlite3.Connection) -> bytes:
    wb = build_unprotected_excel_workbook(db)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
