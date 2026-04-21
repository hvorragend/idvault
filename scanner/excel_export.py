"""
IDV-Scanner Export – SQLite → Excel-Bericht
============================================
Exportiert die Scan-Ergebnisse aus der SQLite-Datenbank
als formatierte Excel-Datei für Revisionen und Fachabteilungen.

Benötigt: pip install openpyxl
"""

import sqlite3
import sys
import argparse
from datetime import datetime, timezone
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:
    print("Bitte openpyxl installieren: pip install openpyxl")
    raise


# ---------------------------------------------------------------------------
# Farben & Styles
# ---------------------------------------------------------------------------

COL_HEADER_BG = "1F497D"   # Dunkelblau (Volksbank-nah)
COL_MACRO_BG  = "FFE599"   # Gelb für Makro-Dateien
COL_CHANGED   = "FCE4D6"   # Orange für geänderte Dateien
COL_NEW       = "E2EFDA"   # Grün für neue Dateien
COL_DELETED   = "F4CCCC"   # Rot für gelöschte Dateien

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def header_style(cell):
    cell.font      = Font(bold=True, color="FFFFFF", size=11)
    cell.fill      = PatternFill("solid", fgColor=COL_HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = BORDER


def data_style(cell, bg=None):
    cell.alignment = Alignment(vertical="center", wrap_text=False)
    cell.border    = BORDER
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)


# ---------------------------------------------------------------------------
# Spalten-Definition
# ---------------------------------------------------------------------------

COLUMNS = [
    ("SHA-256-Hash",           "file_hash",          28),
    ("Dateiname",              "file_name",           35),
    ("Pfad (vollständig)",     "full_path",           60),
    ("Laufwerk/Share",         "share_root",          20),
    ("Relativer Pfad",         "relative_path",       45),
    ("Typ",                    "extension",           8),
    ("Größe (Bytes)",          "size_bytes",          14),
    ("Größe (MB)",             None,                  12),   # berechnet
    ("Erstellt (Dateisystem)", "created_at",          22),
    ("Geändert (Dateisystem)", "modified_at",         22),
    ("Eigentümer",             "file_owner",          25),
    ("Office-Autor",           "office_author",       22),
    ("Zuletzt geändert von",   "office_last_author",  22),
    ("Office-Erstellt",        "office_created",      22),
    ("Office-Geändert",        "office_modified",     22),
    ("Makros (VBA)",           "has_macros",          14),
    ("Externe Verknüpfungen",  "has_external_links",  20),
    ("Tabellenblätter",        "sheet_count",         16),
    ("Benannte Bereiche",      "named_ranges_count",  18),
    ("Erstmals gefunden",      "first_seen_at",       22),
    ("Zuletzt gesehen",        "last_seen_at",        22),
    ("Status",                 "status",              12),
]


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------

def export_to_excel(db_path: str, output_path: str, only_active: bool = True):
    conn = sqlite3.connect(db_path, timeout=15)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode = WAL")
    conn.execute("PRAGMA busy_timeout = 15000")

    # --- Übersicht Scan-Runs ---
    runs = conn.execute(
        "SELECT * FROM scan_runs ORDER BY started_at DESC LIMIT 20"
    ).fetchall()

    # --- Dateiliste ---
    where = "WHERE status = 'active'" if only_active else ""
    files = conn.execute(f"""
        SELECT * FROM idv_files {where}
        ORDER BY extension, file_name
    """).fetchall()

    # --- Delta-Info (letzte Änderungen) ---
    delta = conn.execute("""
            SELECT h.*, f.full_path, f.file_name
            FROM idv_file_history h
            JOIN idv_files f ON h.file_id = f.id
            WHERE h.change_type IN ('new', 'changed', 'deleted')
            ORDER BY h.changed_at DESC
            LIMIT 500
        """).fetchall()

    conn.close()

    wb = openpyxl.Workbook()

    # -----------------------------------------------------------------------
    # Sheet 1: IDV-Grundgesamtheit
    # -----------------------------------------------------------------------
    ws = wb.active
    ws.title = "IDV-Grundgesamtheit"
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 36

    # Header
    for col_idx, (label, _, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        header_style(cell)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Daten
    for row_idx, f in enumerate(files, start=2):
        bg = None
        if f["has_macros"] == 1:
            bg = COL_MACRO_BG

        row_data = []
        for _, field, _ in COLUMNS:
            if field == "has_macros":
                row_data.append("JA" if f["has_macros"] else "nein")
            elif field == "has_external_links":
                row_data.append("JA" if f["has_external_links"] else "nein")
            elif field is None:
                # Größe in MB
                mb = round(f["size_bytes"] / 1024 / 1024, 3) if f["size_bytes"] else None
                row_data.append(mb)
            else:
                row_data.append(f[field] if field in f.keys() else None)

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            data_style(cell, bg=bg)

    # Tabelle / AutoFilter
    if files:
        last_col = get_column_letter(len(COLUMNS))
        last_row = len(files) + 1
        tab = Table(displayName="IDV_Grundgesamtheit",
                    ref=f"A1:{last_col}{last_row}")
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9", showRowStripes=True
        )
        ws.add_table(tab)

    # -----------------------------------------------------------------------
    # Sheet 2: Scan-Zusammenfassung
    # -----------------------------------------------------------------------
    ws2 = wb.create_sheet("Scan-Übersicht")
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 18
    ws2.column_dimensions["D"].width = 18
    ws2.column_dimensions["E"].width = 18
    ws2.column_dimensions["F"].width = 18
    ws2.column_dimensions["G"].width = 18

    headers2 = ["Scan-Run #", "Gestartet", "Beendet", "Gesamt", "Neu", "Geändert", "Gelöscht"]
    for col_idx, h in enumerate(headers2, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        header_style(cell)

    for row_idx, r in enumerate(runs, start=2):
        ws2.cell(row=row_idx, column=1, value=r["id"])
        ws2.cell(row=row_idx, column=2, value=r["started_at"])
        ws2.cell(row=row_idx, column=3, value=r["finished_at"])
        ws2.cell(row=row_idx, column=4, value=r["total_files"])
        cell_new     = ws2.cell(row=row_idx, column=5, value=r["new_files"])
        cell_changed = ws2.cell(row=row_idx, column=6, value=r["changed_files"])
        cell_deleted = ws2.cell(row=row_idx, column=7, value=r["archived_files"])
        if r["new_files"]:      cell_new.fill     = PatternFill("solid", fgColor=COL_NEW)
        if r["changed_files"]:  cell_changed.fill = PatternFill("solid", fgColor=COL_CHANGED)
        if r["archived_files"]: cell_deleted.fill = PatternFill("solid", fgColor=COL_DELETED)

    # -----------------------------------------------------------------------
    # Sheet 3: Delta-Bericht
    # -----------------------------------------------------------------------
    ws3 = wb.create_sheet("Änderungen")
    ws3.freeze_panes = "A2"
    delta_headers = ["Datum", "Änderungstyp", "Dateiname", "Pfad", "Alter Hash", "Neuer Hash"]
    delta_widths  = [22, 14, 35, 60, 28, 28]
    for col_idx, (h, w) in enumerate(zip(delta_headers, delta_widths), start=1):
        cell = ws3.cell(row=1, column=col_idx, value=h)
        header_style(cell)
        ws3.column_dimensions[get_column_letter(col_idx)].width = w

    change_colors = {"new": COL_NEW, "changed": COL_CHANGED, "deleted": COL_DELETED}
    for row_idx, d in enumerate(delta, start=2):
        bg = change_colors.get(d["change_type"])
        keys = d.keys()
        values = [
            d["changed_at"], d["change_type"], d["file_name"],
            d["full_path"],
            d["old_hash"] if "old_hash" in keys else None,
            d["new_hash"] if "new_hash" in keys else None,
        ]
        for col_idx, v in enumerate(values, start=1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=v)
            data_style(cell, bg=bg)

    # -----------------------------------------------------------------------
    # Sheet 4: Statistik-Übersicht
    # -----------------------------------------------------------------------
    ws4 = wb.create_sheet("Statistik")
    now_str = datetime.now(timezone.utc).strftime("%d.%m.%Y %H:%M UTC")
    ws4.column_dimensions["A"].width = 35
    ws4.column_dimensions["B"].width = 20

    stats_data = [
        ("IDV-Scanner Statistik", ""),
        ("Exportiert am", now_str),
        ("", ""),
        ("GESAMTÜBERSICHT", ""),
        ("Dateien aktiv", len([f for f in files if f["status"] == "active"])),
        ("Dateien archiviert (markiert)", len([f for f in files if f["status"] == "archiviert"])),
        ("", ""),
        ("NACH DATEITYP", ""),
    ]

    # Typen zählen
    ext_counts = {}
    for f in files:
        e = f["extension"] or "(unbekannt)"
        ext_counts[e] = ext_counts.get(e, 0) + 1
    for ext, count in sorted(ext_counts.items(), key=lambda x: -x[1]):
        stats_data.append((ext, count))

    stats_data += [
        ("", ""),
        ("MIT VBA-MAKROS", len([f for f in files if f["has_macros"] == 1])),
        ("MIT EXTERNEN VERKNÜPFUNGEN", len([f for f in files if f["has_external_links"] == 1])),
    ]

    for row_idx, (label, value) in enumerate(stats_data, start=1):
        cell_a = ws4.cell(row=row_idx, column=1, value=label)
        cell_b = ws4.cell(row=row_idx, column=2, value=value)
        if label in ("IDV-Scanner Statistik", "GESAMTÜBERSICHT",
                     "NACH DATEITYP", "MIT VBA-MAKROS", "MIT EXTERNEN VERKNÜPFUNGEN"):
            cell_a.font = Font(bold=True, size=12)

    # Speichern
    wb.save(output_path)
    print(f"Export gespeichert: {output_path}")
    print(f"  Dateien exportiert  : {len(files)}")
    print(f"  Scan-Runs enthalten : {len(runs)}")
    print(f"  Änderungen enthalten: {len(delta)}")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="IDV-Scanner Excel-Export")
    parser.add_argument("--db",     default="idv_register.db", help="Pfad zur SQLite-DB")
    parser.add_argument("--output", default=None,              help="Ausgabe-Excel-Datei")
    parser.add_argument("--all",    action="store_true",       help="Auch gelöschte Dateien exportieren")
    args = parser.parse_args()

    output = args.output or f"IDV_Grundgesamtheit_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    export_to_excel(args.db, output, only_active=not args.all)
